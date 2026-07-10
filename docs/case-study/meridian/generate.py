#!/usr/bin/env python3
"""
Meridian SaaS — exec-metrics case study generator (design-brief 01).

Produces a coherent, internally-consistent synthetic B2B-SaaS reporting workbook:
raw data (subscriptions, invoices, FX with a deliberate gap, targets) plus the
DERIVED sheets (MRR movement waterfall, cohort-retention triangle, top customers,
exec summary), all computed here in Python so the numbers are real and the
reconciliation invariants actually hold (the MRR waterfall sums to the MRR delta;
cohorts are genuine). Emits CSVs + a real .xlsx carrying live Excel formulas.

Deterministic: fixed seed. Re-running yields the identical workbook.

Usage:  python3 generate.py        # writes data/*.csv, meridian.xlsx, expected.json
"""
import csv, json, os, random
from collections import defaultdict, OrderedDict
from datetime import date

SEED = 20260709
random.seed(SEED)

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
os.makedirs(DATA, exist_ok=True)

# ── calendar: 36 months, 2023-07 .. 2026-06 ──────────────────────────────────
def months(start_y, start_m, n):
    out = []
    y, m = start_y, start_m
    for _ in range(n):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1; y += 1
    return out

MONTHS = months(2023, 7, 36)
MONTH_IDX = {m: i for i, m in enumerate(MONTHS)}

def add_months(mkey, k):
    i = MONTH_IDX[mkey] + k
    return MONTHS[i] if 0 <= i < len(MONTHS) else None

# ── reference data ───────────────────────────────────────────────────────────
PLANS = {  # plan -> (per-seat monthly price in the customer's currency-units baseline USD)
    "Starter":    30,
    "Growth":     90,
    "Scale":      240,
    "Enterprise": 600,
}
PLAN_ORDER = ["Starter", "Growth", "Scale", "Enterprise"]
CURRENCIES = ["USD", "EUR", "GBP", "CAD"]
REGIONS = {"USD": "amer", "EUR": "emea", "GBP": "emea", "CAD": "amer"}

# FX: units of currency per 1 EUR (so EUR amount = local / rate). Base = EUR (1.0).
# A DELIBERATE GAP at 2024-02 for GBP exercises as-of / carry-forward.
def fx_series():
    rows = []
    base = {"USD": 1.08, "EUR": 1.00, "GBP": 0.86, "CAD": 1.47}
    for mi, mk in enumerate(MONTHS):
        for cur in CURRENCIES:
            if cur == "GBP" and mk == "2024-02":
                continue  # the gap: no GBP rate this month -> carry-forward last known
            drift = 1 + 0.015 * random.uniform(-1, 1) + 0.0008 * mi * (1 if cur == "USD" else -1 if cur == "CAD" else 0)
            rate = round(base[cur] * drift, 4) if cur != "EUR" else 1.0
            rows.append({"month": mk, "currency": cur, "eur_per_unit_inv": rate})
    return rows

FX = fx_series()
# resolve an as-of (carry-forward) rate map for our own EUR normalization
_fx_lookup = defaultdict(dict)
for r in FX:
    _fx_lookup[r["currency"]][r["month"]] = r["eur_per_unit_inv"]
def rate_asof(cur, mk):
    if cur == "EUR":
        return 1.0
    i = MONTH_IDX[mk]
    while i >= 0:
        mm = MONTHS[i]
        if mm in _fx_lookup[cur]:
            return _fx_lookup[cur][mm]
        i -= 1
    return _fx_lookup[cur][MONTHS[0]]  # fallback

def to_eur(amount_local, cur, mk):
    return amount_local / rate_asof(cur, mk)

# ── customers & their monthly subscription lifecycle ─────────────────────────
N_CUST = 150
customers = []
_cust_cur = {}  # customer_id -> currency (filled below; used by EUR normalization)
for cid in range(1, N_CUST + 1):
    cur = random.choices(CURRENCIES, weights=[45, 30, 15, 10])[0]
    signup_i = random.randint(0, 32)  # leave room to observe
    plan = random.choices(PLAN_ORDER, weights=[40, 35, 18, 7])[0]
    seats = random.choice([3, 5, 8, 10, 12, 20, 25, 40, 60, 100])
    # lifetime + churn hazard
    life = random.randint(4, 40)
    customers.append({
        "customer_id": f"C{cid:04d}",
        "name": f"Customer {cid:04d}",
        "currency": cur, "region": REGIONS[cur],
        "signup_i": signup_i, "plan": plan, "seats": seats, "life": life,
    })
    _cust_cur[f"C{cid:04d}"] = cur

# Build one row per customer-month while active, applying expansion / contraction /
# churn / reactivation events so the movement waterfall has real signal.
sub_rows = []            # subscriptions.csv rows
state_by_month = defaultdict(dict)   # month -> {cust_id: mrr_local} for movement calc
meta_by_cust = {}        # cust_id -> current plan/seats/currency (mutated over time)

for c in customers:
    cid = c["customer_id"]
    active = False
    plan, seats = c["plan"], c["seats"]
    reactivated_once = False
    end_i = min(c["signup_i"] + c["life"], 36)
    for mi in range(c["signup_i"], end_i):
        mk = MONTHS[mi]
        # random mid-life events
        roll = random.random()
        if active:
            if roll < 0.05 and seats > 3:            # contraction
                seats = max(3, seats - random.choice([1, 2, 3]))
            elif roll < 0.12:                         # expansion (seats and/or plan up)
                seats += random.choice([1, 2, 3, 5])
                if random.random() < 0.3 and PLAN_ORDER.index(plan) < 3:
                    plan = PLAN_ORDER[PLAN_ORDER.index(plan) + 1]
        active = True
        mrr_local = seats * PLANS[plan]
        state_by_month[mk][cid] = mrr_local
        sub_rows.append(OrderedDict([
            ("month", mk), ("customer_id", cid), ("name", c["name"]),
            ("region", c["region"]), ("currency", c["currency"]),
            ("plan", plan), ("seats", seats),
            ("mrr_local", mrr_local),
            ("signup_month", MONTHS[c["signup_i"]]),
        ]))
    # a fraction reactivate a few months after churn
    if end_i < 30 and random.random() < 0.15:
        gap = random.randint(2, 4)
        re_i = end_i + gap
        if re_i <= 32:                       # leave room for >=3 reactivated months
            plan2, seats2 = c["plan"], max(3, c["seats"] - 1)
            life2 = random.randint(3, 36 - re_i)
            for mi in range(re_i, min(re_i + life2, 36)):
                mk = MONTHS[mi]
                seats2 += 1 if random.random() < 0.1 else 0
                mrr_local = seats2 * PLANS[plan2]
                state_by_month[mk][cid] = mrr_local
                sub_rows.append(OrderedDict([
                    ("month", mk), ("customer_id", cid), ("name", c["name"]),
                    ("region", c["region"]), ("currency", c["currency"]),
                    ("plan", plan2), ("seats", seats2),
                    ("mrr_local", mrr_local), ("signup_month", MONTHS[c["signup_i"]]),
                ]))

sub_rows.sort(key=lambda r: (r["month"], r["customer_id"]))

# ── invoices (one per active customer-month, in the customer's currency) ─────
inv_rows = []
inv_no = 1000
for r in sub_rows:
    inv_no += 1
    inv_rows.append(OrderedDict([
        ("invoice_id", f"INV{inv_no}"), ("month", r["month"]),
        ("customer_id", r["customer_id"]), ("currency", r["currency"]),
        ("amount_local", r["mrr_local"]),
    ]))

# ── first-seen month per customer (for new-vs-reactivation split) ────────────
first_seen = {}
for r in sub_rows:
    first_seen.setdefault(r["customer_id"], r["month"])

# ── MRR movement waterfall (computed so it RECONCILES to ΔMRR) ────────────────
# EUR-normalized. Categories: new, expansion, contraction, churned, reactivation.
# new = first-ever-seen this month; reactivation = seen before, absent last month.
mov_rows = []
prev_state = {}
for mi, mk in enumerate(MONTHS):
    cur_state = state_by_month.get(mk, {})
    cur_eur = {c: to_eur(v, _cust_cur[c], mk) for c, v in cur_state.items()}
    prev_eur = {c: to_eur(v, _cust_cur[c], MONTHS[mi-1]) for c, v in prev_state.items()} if mi else {}
    new = expa = cont = chur = react = 0.0
    for c, v in cur_eur.items():
        if c not in prev_eur:
            if first_seen[c] == mk: new += v
            else: react += v
        else:
            d = v - prev_eur[c]
            if d > 0: expa += d
            elif d < 0: cont += d               # negative
    for c, v in prev_eur.items():
        if c not in cur_eur:
            chur -= v                           # negative
    start_mrr = sum(prev_eur.values()); end_mrr = sum(cur_eur.values())
    mov_rows.append(OrderedDict([
        ("month", mk), ("start_mrr", round(start_mrr, 2)),
        ("new", round(new, 2)), ("expansion", round(expa, 2)),
        ("contraction", round(cont, 2)), ("churned", round(chur, 2)),
        ("reactivation", round(react, 2)), ("end_mrr", round(end_mrr, 2)),
        # reconciliation: start + all movements - end  (must be ~0 -> conservation)
        ("recon_delta", round((start_mrr+new+expa+cont+chur+react)-end_mrr, 2)),
    ]))
    prev_state = cur_state

# ── cohort retention triangle (signup cohort × months-since) ─────────────────
# active customers per (cohort, offset), and retention % vs offset 0.
cohort_active = defaultdict(lambda: defaultdict(set))  # cohort -> offset -> {cust}
for r in sub_rows:
    cohort = r["signup_month"]
    off = MONTH_IDX[r["month"]] - MONTH_IDX[cohort]
    if off >= 0:
        cohort_active[cohort][off].add(r["customer_id"])
cohorts = sorted(cohort_active.keys())
max_off = 36
coh_rows = []
for cohort in cohorts:
    size0 = len(cohort_active[cohort].get(0, set()))
    row = OrderedDict([("cohort", cohort), ("cohort_size", size0)])
    for off in range(0, 13):  # first 12 months since signup
        cnt = len(cohort_active[cohort].get(off, set()))
        # only meaningful where the offset month exists in our window
        if MONTH_IDX[cohort] + off < 36 and size0 > 0:
            row[f"m{off}_pct"] = round(100.0 * cnt / size0, 1)
        else:
            row[f"m{off}_pct"] = ""
    coh_rows.append(row)

# ── top customers by ARR (latest month) + Other bucket + concentration ───────
latest = MONTHS[-1]
arr_by_cust = {}
for c, v in state_by_month.get(latest, {}).items():
    arr_by_cust[c] = round(to_eur(v, _cust_cur[c], latest) * 12, 2)
ranked = sorted(arr_by_cust.items(), key=lambda kv: kv[1], reverse=True)
total_arr = sum(arr_by_cust.values())
top_rows = []
top10 = ranked[:10]
for rank, (cid, arr) in enumerate(top10, 1):
    top_rows.append(OrderedDict([
        ("rank", rank), ("customer_id", cid),
        ("arr_eur", arr), ("share_pct", round(100*arr/total_arr, 1)),
    ]))
other_arr = round(sum(a for _, a in ranked[10:]), 2)
top_rows.append(OrderedDict([
    ("rank", ""), ("customer_id", "Other"),
    ("arr_eur", other_arr), ("share_pct", round(100*other_arr/total_arr, 1)),
]))
top10_concentration = round(100 * sum(a for _, a in top10) / total_arr, 1)

# ── targets (base / bull / bear) ─────────────────────────────────────────────
tgt_rows = []
base_start = mov_rows[6]["end_mrr"] if len(mov_rows) > 6 else 50000
for mi, mk in enumerate(MONTHS):
    g = 1 + 0.03 * mi
    tgt_rows.append(OrderedDict([
        ("month", mk),
        ("mrr_base", round(base_start * g, 0)),
        ("mrr_bull", round(base_start * (1 + 0.045 * mi), 0)),
        ("mrr_bear", round(base_start * (1 + 0.018 * mi), 0)),
        ("nrr_target_pct", 108),
    ]))

# ── exec summary (KPIs per month, EUR) ───────────────────────────────────────
def safe_div(a, b): return round(a / b, 4) if b else None
exec_rows = []
for mi, mk in enumerate(MONTHS):
    m = mov_rows[mi]
    start = m["start_mrr"]; end = m["end_mrr"]
    # NRR = (start + expansion + contraction + churn) / start   (excludes new)
    nrr = safe_div(start + m["expansion"] + m["contraction"] + m["churned"], start)
    grr = safe_div(start + m["contraction"] + m["churned"], start)
    quick_ratio = safe_div(m["new"] + m["expansion"], -(m["contraction"] + m["churned"]))
    growth = safe_div(end - start, start)
    churn_pct = safe_div(-(m["churned"]), start)
    exec_rows.append(OrderedDict([
        ("month", mk), ("mrr", end), ("mom_growth_pct", None if growth is None else round(growth*100,1)),
        ("nrr_pct", None if nrr is None else round(nrr*100,1)),
        ("grr_pct", None if grr is None else round(grr*100,1)),
        ("quick_ratio", quick_ratio),
        ("gross_churn_pct", None if churn_pct is None else round(churn_pct*100,1)),
        ("vs_target_base_pct", safe_div(end, tgt_rows[mi]["mrr_base"])),
    ]))

# ── write CSVs ───────────────────────────────────────────────────────────────
def write_csv(name, rows, fieldnames=None):
    if not rows: return
    fieldnames = fieldnames or list(rows[0].keys())
    with open(os.path.join(DATA, name), "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader(); w.writerows(rows)

write_csv("subscriptions.csv", sub_rows)
write_csv("invoices.csv", inv_rows)
write_csv("fx.csv", FX)
write_csv("targets.csv", tgt_rows)
write_csv("mrr_movements.csv", mov_rows)
write_csv("cohort_retention.csv", coh_rows)
write_csv("top_customers.csv", top_rows)
write_csv("exec_summary.csv", exec_rows)

# ── expected-values spot-check file (seeds the Reckoner `specification` tests) ─
max_recon = max(abs(m["recon_delta"]) for m in mov_rows)
expected = {
    "seed": SEED,
    "months": len(MONTHS),
    "subscription_rows": len(sub_rows),
    "customers": N_CUST,
    "latest_month": latest,
    "latest_mrr_eur": exec_rows[-1]["mrr"],
    "latest_nrr_pct": exec_rows[-1]["nrr_pct"],
    "top10_concentration_pct": top10_concentration,
    "other_arr_eur": other_arr,
    "waterfall_max_abs_recon_delta": max_recon,   # MUST be ~0 -> the conservation invariant
    "fx_gap": {"currency": "GBP", "missing_month": "2024-02", "carry_forward_from": "2024-01"},
    "spot_checks": {
        "mov_2024_06": mov_rows[11] if len(mov_rows) > 11 else None,
        "exec_2025_12": next((e for e in exec_rows if e["month"] == "2025-12"), None),
    },
}
with open(os.path.join(HERE, "expected.json"), "w") as f:
    json.dump(expected, f, indent=2)

# ── build the .xlsx with LIVE Excel formulas (the "before" of the port) ───────
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    def sheet_from_rows(title, rows, first=False):
        ws = wb.active if first else wb.create_sheet(title)
        if first: ws.title = title
        if not rows: return ws
        cols = list(rows[0].keys())
        ws.append(cols)
        for r in rows:
            ws.append([r[c] for c in cols])
        return ws

    sheet_from_rows("Subscriptions", sub_rows, first=True)
    sheet_from_rows("Invoices", inv_rows)
    sheet_from_rows("FX", FX)
    sheet_from_rows("Targets", tgt_rows)

    # MRR movements: keep computed values AND add a live reconciliation formula column
    ws = sheet_from_rows("MRR movements", mov_rows)
    recon_col = len(mov_rows[0]) + 1
    ws.cell(row=1, column=recon_col, value="recon_formula")
    # start + new + expansion + contraction + churned + reactivation - end  (Excel, live)
    hdr = list(mov_rows[0].keys())
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    for ri in range(2, len(mov_rows) + 2):
        def L(h): return f"{get_column_letter(ci[h])}{ri}"
        ws.cell(row=ri, column=recon_col,
                value=f"={L('start_mrr')}+{L('new')}+{L('expansion')}+{L('contraction')}+{L('churned')}+{L('reactivation')}-{L('end_mrr')}")

    sheet_from_rows("Cohort retention", coh_rows)
    ws_top = sheet_from_rows("Top customers", top_rows)
    ws_top.cell(row=len(top_rows) + 3, column=1, value="top10_concentration_pct")
    ws_top.cell(row=len(top_rows) + 3, column=2, value=top10_concentration)
    sheet_from_rows("Exec summary", exec_rows)

    wb.save(os.path.join(HERE, "meridian.xlsx"))
    xlsx_status = "meridian.xlsx written (with live recon formulas)"
except Exception as e:  # openpyxl absent -> CSV-only is fine (brief 01 allows CSV)
    xlsx_status = f"xlsx skipped ({e}); CSVs are the artifact"

print(f"subscriptions rows : {len(sub_rows)}")
print(f"customers          : {N_CUST}")
print(f"months             : {len(MONTHS)} ({MONTHS[0]}..{MONTHS[-1]})")
print(f"waterfall max |recon delta| : {max_recon}  (must be ~0 -> conservation holds)")
print(f"top-10 concentration : {top10_concentration}%")
print(f"latest MRR (EUR)   : {exec_rows[-1]['mrr']}  NRR {exec_rows[-1]['nrr_pct']}%")
print(f"FX gap             : GBP 2024-02 missing (carry-forward)")
print(xlsx_status)
